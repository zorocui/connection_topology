from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import Cluster, Device, ImportStatus, ImportTestStatus, OSType
from app.services.imports import (
    IMPORT_HEADERS,
    ImportValidationError,
    build_import_report,
    build_import_template,
    import_devices,
)


def workbook_bytes(rows) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "设备导入"
    sheet.append(IMPORT_HEADERS)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_template_has_required_sheets_and_headers():
    workbook = load_workbook(BytesIO(build_import_template()), data_only=True)
    assert workbook.sheetnames == ["设备导入", "填写说明"]
    assert tuple(cell.value for cell in workbook["设备导入"][1]) == IMPORT_HEADERS
    instructions = {
        row[0]: row[2]
        for row in workbook["填写说明"].iter_rows(
            min_row=2,
            values_only=True,
        )
    }
    assert "新集群" in instructions["采集间隔（分钟）"]
    assert "已有集群" in instructions["采集间隔（分钟）"]
    assert "可留空" in instructions["密码"]
    assert "所属集群" in instructions["密码"]


def test_passwordless_row_requires_cluster(app):
    content = workbook_bytes(
        [("marker", "10.0.0.70", "linux", 22, "ops", "", "", 5, "否")]
    )
    with app.state.session_factory() as session:
        batch = import_devices(session, app.state.cipher, "devices.xlsx", content)
        assert batch.error_rows == 1
        assert batch.imported_rows == 0
        assert batch.rows[0].import_message == "未填写密码时必须填写所属集群"


def test_passwordless_cluster_row_creates_non_collectible_member(app):
    content = workbook_bytes(
        [
            (
                "marker",
                "10.0.0.71",
                "linux",
                22,
                "ops",
                "",
                "marker-cluster",
                5,
                "是",
            )
        ]
    )
    with app.state.session_factory() as session:
        batch = import_devices(session, app.state.cipher, "devices.xlsx", content)
        device = session.scalar(select(Device).where(Device.host == "10.0.0.71"))
        assert batch.status.value == "completed"
        assert batch.imported_rows == 1
        assert batch.test_pending_rows == 0
        assert device.collection_enabled is False
        assert app.state.cipher.decrypt(device.encrypted_password) == ""
        assert device.cluster.name == "marker-cluster"
        assert batch.rows[0].test_status == ImportTestStatus.NOT_APPLICABLE
        assert batch.rows[0].import_message == "仅标注集群设备，不执行连接测试"


def test_mixed_import_tests_only_password_devices(app):
    content = workbook_bytes(
        [
            ("real", "10.0.0.72", "linux", 22, "ops", "secret", "mixed", 5, "是"),
            ("marker", "10.0.0.73", "linux", 22, "ops", "", "mixed", 5, "是"),
        ]
    )
    with app.state.session_factory() as session:
        batch = import_devices(session, app.state.cipher, "devices.xlsx", content)
        assert batch.test_pending_rows == 1
        assert batch.status.value == "testing"
        statuses = {row.device_name: row.test_status for row in batch.rows}
        assert statuses == {
            "real": ImportTestStatus.PENDING,
            "marker": ImportTestStatus.NOT_APPLICABLE,
        }


def test_passwordless_duplicate_is_logged_and_does_not_modify_device(app, caplog):
    content = workbook_bytes(
        [
            (
                "marker-copy",
                "10.0.0.74",
                "linux",
                22,
                "ops",
                "",
                "new-cluster",
                5,
                "否",
            )
        ]
    )
    with app.state.session_factory() as session:
        original = Device(
            name="original",
            host="10.0.0.74",
            os_type=OSType.LINUX,
            port=22,
            username="ops",
            encrypted_password=app.state.cipher.encrypt("secret"),
        )
        session.add(original)
        session.commit()
        original_id = original.id

        with caplog.at_level("WARNING", logger="app.services.imports"):
            batch = import_devices(session, app.state.cipher, "devices.xlsx", content)

        session.refresh(original)
        assert batch.error_rows == 1
        assert batch.skipped_rows == 0
        assert batch.rows[0].import_status == ImportStatus.ERROR
        assert batch.rows[0].import_message == "设备已存在，未修改集群归属"
        assert batch.rows[0].device_id == original_id
        assert original.cluster_id is None
        assert original.name == "original"
        for expected in (
            f"batch_id={batch.id}",
            "row=2",
            "name=marker-copy",
            "host=10.0.0.74",
            "port=22",
            "username=ops",
            "设备已存在，未修改集群归属",
        ):
            assert expected in caplog.text
        assert "secret" not in caplog.text
        assert original.encrypted_password not in caplog.text


def test_mixed_import_encrypts_password_and_creates_cluster(app):
    content = workbook_bytes(
        [
            ("web-01", "10.0.0.10", "linux", None, "ops", "secret-1", "生产集群", None, "是"),
            ("bad", "10.0.0.11", "other", 22, "ops", "secret-2", "", 5, "是"),
            ("web-duplicate", "10.0.0.10", "linux", 22, "ops", "secret-3", "", 5, "否"),
        ]
    )
    with app.state.session_factory() as session:
        batch = import_devices(session, app.state.cipher, "devices.xlsx", content)
        assert batch.total_rows == 3
        assert batch.imported_rows == 1
        assert batch.error_rows == 1
        assert batch.skipped_rows == 1
        assert batch.test_running_rows == 0
        device = session.scalar(select(Device).where(Device.host == "10.0.0.10"))
        assert device.port == 22
        assert "secret-1" not in device.encrypted_password
        assert app.state.cipher.decrypt(device.encrypted_password) == "secret-1"
        imported_cluster = session.scalar(
            select(Cluster).where(Cluster.name == "生产集群")
        )
        assert imported_cluster is not None
        assert imported_cluster.internal_networks == []
        rows = batch.rows
        assert {row.import_status for row in rows} == {
            ImportStatus.IMPORTED,
            ImportStatus.ERROR,
            ImportStatus.SKIPPED,
        }
        imported = next(row for row in rows if row.import_status == ImportStatus.IMPORTED)
        assert imported.test_status == ImportTestStatus.PENDING
        report = build_import_report(session, batch.id)
        assert b"secret-1" not in report
        report_book = load_workbook(BytesIO(report), data_only=True)
        assert report_book["导入结果"].max_row == 4

        imported.test_status = ImportTestStatus.RUNNING
        imported.test_message = "正在测试连接"
        batch.test_pending_rows = 0
        batch.test_running_rows = 1
        session.commit()
        running_report = load_workbook(
            BytesIO(build_import_report(session, batch.id)),
            data_only=True,
        )
        report_rows = list(running_report["导入结果"].values)
        imported_report_row = next(row for row in report_rows if row[1] == "web-01")
        assert imported_report_row[5:] == ("正在测试", "正在测试连接")


def test_import_rejects_non_xlsx(app):
    with app.state.session_factory() as session:
        try:
            import_devices(session, app.state.cipher, "devices.csv", b"x")
        except ImportValidationError as exc:
            assert ".xlsx" in str(exc)
        else:
            raise AssertionError("应拒绝非 xlsx 文件")


def test_imported_device_uses_existing_cluster_scan_policy(app):
    content = workbook_bytes(
        [
            (
                "worker",
                "10.0.0.50",
                "linux",
                22,
                "ops",
                "secret",
                "policy-cluster",
                2,
                "是",
            )
        ]
    )
    with app.state.session_factory() as session:
        session.add(
            Cluster(
                name="policy-cluster",
                scan_interval_minutes=17,
                scheduled_enabled=False,
            )
        )
        session.commit()

        import_devices(
            session,
            app.state.cipher,
            "devices.xlsx",
            content,
        )
        device = session.scalar(
            select(Device).where(Device.host == "10.0.0.50")
        )

        assert device.scan_interval_minutes == 17
        assert device.scheduled_enabled is False


def test_import_creates_cluster_with_row_scan_policy(app):
    content = workbook_bytes(
        [
            (
                "new-worker",
                "10.0.0.60",
                "linux",
                22,
                "ops",
                "secret",
                "new-policy-cluster",
                60,
                "否",
            )
        ]
    )

    with app.state.session_factory() as session:
        import_devices(session, app.state.cipher, "devices.xlsx", content)
        cluster = session.scalar(
            select(Cluster).where(Cluster.name == "new-policy-cluster")
        )
        device = session.scalar(
            select(Device).where(Device.host == "10.0.0.60")
        )

        assert cluster.scan_interval_minutes == 60
        assert cluster.scheduled_enabled is False
        assert device.scan_interval_minutes == 60
        assert device.scheduled_enabled is False


def test_import_uses_first_row_policy_for_same_new_cluster(app):
    content = workbook_bytes(
        [
            (
                "first-worker",
                "10.0.0.61",
                "linux",
                22,
                "ops",
                "secret",
                "shared-new-cluster",
                60,
                "否",
            ),
            (
                "second-worker",
                "10.0.0.62",
                "linux",
                22,
                "ops",
                "secret",
                "shared-new-cluster",
                10,
                "是",
            ),
        ]
    )

    with app.state.session_factory() as session:
        import_devices(session, app.state.cipher, "devices.xlsx", content)
        cluster = session.scalar(
            select(Cluster).where(Cluster.name == "shared-new-cluster")
        )
        devices = session.scalars(
            select(Device).where(Device.cluster_id == cluster.id)
        ).all()

        assert cluster.scan_interval_minutes == 60
        assert cluster.scheduled_enabled is False
        assert len(devices) == 2
        assert {device.scan_interval_minutes for device in devices} == {60}
        assert {device.scheduled_enabled for device in devices} == {False}


def test_template_download_api(client):
    response = client.get("/api/imports/template")
    assert response.status_code == 200
    assert response.content[:2] == b"PK"
