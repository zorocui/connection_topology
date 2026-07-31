from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.models import (
    Device,
    ImportBatch,
    ImportBatchStatus,
    ImportRowResult,
    ImportStatus,
    ImportTestStatus,
    OSType,
)
from app.security import CredentialCipher
from app.services.clusters import (
    cluster_scan_values,
    create_cluster,
    find_cluster_by_name,
)

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
IMPORT_SHEET = "设备导入"
IMPORT_HEADERS = (
    "设备名称",
    "主机地址",
    "操作系统",
    "端口",
    "用户名",
    "密码",
    "所属集群",
    "采集间隔（分钟）",
    "启用定时采集",
)


class ImportValidationError(ValueError):
    pass


def _style_header(sheet, cell_range: str) -> None:
    header = sheet[cell_range][0]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor="163239")
        cell.font = Font(color="A6FFCB", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="4B6B72"))


def build_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = IMPORT_SHEET
    sheet.append(IMPORT_HEADERS)
    sheet.append(
        (
            "示例-Web-01",
            "10.0.0.10",
            "linux",
            22,
            "ops",
            "请替换为真实密码",
            "生产集群",
            5,
            "是",
        )
    )
    _style_header(sheet, "A1:I1")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:I2"
    widths = (20, 24, 14, 10, 18, 24, 20, 18, 18)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False
    os_validation = DataValidation(type="list", formula1='"linux,windows"')
    enabled_validation = DataValidation(type="list", formula1='"是,否"')
    sheet.add_data_validation(os_validation)
    sheet.add_data_validation(enabled_validation)
    os_validation.add("C2:C1001")
    enabled_validation.add("I2:I1001")

    guide = workbook.create_sheet("填写说明")
    guide.sheet_view.showGridLines = False
    guide.append(("字段", "是否必填", "填写规则"))
    instructions = (
        ("设备名称", "是", "用于页面识别，最长 100 个字符"),
        ("主机地址", "是", "IP 地址或可解析的主机名"),
        ("操作系统", "是", "仅填写 linux 或 windows"),
        ("端口", "否", "Linux 默认 22，Windows 默认 5985"),
        ("用户名", "是", "远程登录用户名"),
        ("密码", "是", "明文仅用于导入；导入后请删除或加密保存源文件"),
        ("所属集群", "否", "不存在时自动创建，留空表示未分组"),
        (
            "采集间隔（分钟）",
            "否",
            "未分组设备直接采用；新集群以首条成功记录为准；已有集群继承集群设置",
        ),
        (
            "启用定时采集",
            "否",
            "填写 是 或 否；未分组设备直接采用；新集群以首条成功记录为准；已有集群继承集群设置",
        ),
    )
    for row in instructions:
        guide.append(row)
    _style_header(guide, "A1:C1")
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 14
    guide.column_dimensions["C"].width = 58
    for row in guide.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    guide.freeze_panes = "A2"

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _text(value, field: str, *, required: bool = True, max_length: int = 255) -> str:
    text = "" if value is None else str(value).strip()
    if required and not text:
        raise ImportValidationError(f"{field}不能为空")
    if len(text) > max_length:
        raise ImportValidationError(f"{field}不能超过 {max_length} 个字符")
    return text


def _integer(value, field: str, default: int, minimum: int, maximum: int) -> int:
    if value in {None, ""}:
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ImportValidationError(f"{field}必须是整数") from exc
    if not minimum <= parsed <= maximum:
        raise ImportValidationError(f"{field}必须在 {minimum} 到 {maximum} 之间")
    return parsed


def _enabled(value) -> bool:
    if value in {None, ""}:
        return True
    normalized = str(value).strip()
    if normalized == "是":
        return True
    if normalized == "否":
        return False
    raise ImportValidationError("启用定时采集只能填写“是”或“否”")


def _parse_row(values: tuple) -> dict:
    name = _text(values[0], "设备名称", max_length=100)
    host = _text(values[1], "主机地址")
    os_value = _text(values[2], "操作系统").lower()
    if os_value not in {"linux", "windows"}:
        raise ImportValidationError("操作系统只能填写 linux 或 windows")
    os_type = OSType(os_value)
    default_port = 22 if os_type == OSType.LINUX else 5985
    return {
        "name": name,
        "host": host,
        "os_type": os_type,
        "port": _integer(values[3], "端口", default_port, 1, 65535),
        "username": _text(values[4], "用户名"),
        "password": _text(values[5], "密码", max_length=1024),
        "cluster_name": _text(values[6], "所属集群", required=False, max_length=100),
        "scan_interval_minutes": _integer(
            values[7], "采集间隔（分钟）", 5, 1, 10080
        ),
        "scheduled_enabled": _enabled(values[8]),
    }


def _load_rows(filename: str, content: bytes) -> list[tuple[int, tuple]]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ImportValidationError("只支持 .xlsx 文件")
    if len(content) > MAX_IMPORT_BYTES:
        raise ImportValidationError("文件不能超过 5 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError("Excel 文件损坏或格式无效") from exc
    if IMPORT_SHEET not in workbook.sheetnames:
        raise ImportValidationError("缺少“设备导入”工作表")
    sheet = workbook[IMPORT_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(next(rows, ()))
    if headers != IMPORT_HEADERS:
        raise ImportValidationError("Excel 表头与模板不一致")
    result = [
        (row_number, tuple(values))
        for row_number, values in enumerate(rows, start=2)
        if any(value not in {None, ""} for value in values)
    ]
    if len(result) > MAX_IMPORT_ROWS:
        raise ImportValidationError("数据行不能超过 1000 行")
    return result


def import_devices(
    session: Session,
    cipher: CredentialCipher,
    filename: str,
    content: bytes,
) -> ImportBatch:
    rows = _load_rows(filename, content)
    batch = ImportBatch(
        filename=Path(filename).name[:255],
        status=ImportBatchStatus.IMPORTING,
        total_rows=len(rows),
    )
    session.add(batch)
    session.commit()
    session.refresh(batch)

    for row_number, values in rows:
        device_name = None
        host = None
        try:
            parsed = _parse_row(values)
            device_name = parsed["name"]
            host = parsed["host"]
            duplicate = session.scalar(
                select(Device).where(
                    Device.host == parsed["host"],
                    Device.port == parsed["port"],
                    Device.username == parsed["username"],
                )
            )
            if duplicate:
                result = ImportRowResult(
                    batch_id=batch.id,
                    row_number=row_number,
                    device_name=device_name,
                    host=host,
                    device_id=duplicate.id,
                    import_status=ImportStatus.SKIPPED,
                    import_message="设备已存在，已跳过",
                    test_status=ImportTestStatus.NOT_APPLICABLE,
                )
                session.add(result)
                batch.skipped_rows += 1
                session.commit()
                continue
            cluster = None
            if parsed["cluster_name"]:
                cluster = find_cluster_by_name(session, parsed["cluster_name"])
                if cluster is None:
                    cluster = create_cluster(
                        session,
                        parsed["cluster_name"],
                        scan_interval_minutes=parsed["scan_interval_minutes"],
                        scheduled_enabled=parsed["scheduled_enabled"],
                    )
            scan_interval, scheduled_enabled = cluster_scan_values(
                cluster,
                parsed["scan_interval_minutes"],
                parsed["scheduled_enabled"],
            )
            device = Device(
                name=parsed["name"],
                host=parsed["host"],
                os_type=parsed["os_type"],
                port=parsed["port"],
                username=parsed["username"],
                encrypted_password=cipher.encrypt(parsed["password"]),
                cluster_id=cluster.id if cluster else None,
                scan_interval_minutes=scan_interval,
                scheduled_enabled=scheduled_enabled,
            )
            session.add(device)
            session.flush()
            session.add(
                ImportRowResult(
                    batch_id=batch.id,
                    row_number=row_number,
                    device_name=device_name,
                    host=host,
                    device_id=device.id,
                    import_status=ImportStatus.IMPORTED,
                    import_message="导入成功，等待连接测试",
                    test_status=ImportTestStatus.PENDING,
                )
            )
            batch.imported_rows += 1
            batch.test_pending_rows += 1
            session.commit()
        except Exception as exc:  # noqa: BLE001 - every workbook row is isolated
            session.rollback()
            batch = session.get(ImportBatch, batch.id)
            assert batch is not None
            batch.error_rows += 1
            if isinstance(exc, ImportValidationError):
                message = str(exc)
            elif isinstance(exc, IntegrityError):
                message = "数据库写入冲突"
            else:
                message = "该行导入失败"
            session.add(
                ImportRowResult(
                    batch_id=batch.id,
                    row_number=row_number,
                    device_name=device_name,
                    host=host,
                    import_status=ImportStatus.ERROR,
                    import_message=message,
                    test_status=ImportTestStatus.NOT_APPLICABLE,
                )
            )
            session.commit()
    batch = session.get(ImportBatch, batch.id)
    assert batch is not None
    batch.status = (
        ImportBatchStatus.TESTING
        if batch.test_pending_rows
        else ImportBatchStatus.COMPLETED
    )
    if batch.status == ImportBatchStatus.COMPLETED:
        batch.finished_at = datetime.now(timezone.utc)
    session.commit()
    session.refresh(batch)
    return batch


def build_import_report(session: Session, batch_id: int) -> bytes:
    batch = session.scalar(
        select(ImportBatch)
        .where(ImportBatch.id == batch_id)
        .options(selectinload(ImportBatch.rows))
    )
    if batch is None:
        raise ImportValidationError("导入批次不存在")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入结果"
    sheet.append(
        (
            "行号",
            "设备名称",
            "主机地址",
            "导入状态",
            "导入说明",
            "连接测试状态",
            "连接测试说明",
        )
    )
    status_names = {
        ImportStatus.IMPORTED: "已导入",
        ImportStatus.SKIPPED: "已跳过",
        ImportStatus.ERROR: "错误",
    }
    test_names = {
        ImportTestStatus.PENDING: "待测试",
        ImportTestStatus.SUCCESS: "成功",
        ImportTestStatus.FAILED: "失败",
        ImportTestStatus.NOT_APPLICABLE: "不适用",
    }
    for row in sorted(batch.rows, key=lambda item: item.row_number):
        sheet.append(
            (
                row.row_number,
                row.device_name,
                row.host,
                status_names[row.import_status],
                row.import_message,
                test_names[row.test_status],
                row.test_message,
            )
        )
    _style_header(sheet, "A1:G1")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(1, sheet.max_row)}"
    widths = (10, 22, 24, 14, 36, 18, 40)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.sheet_view.showGridLines = False
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
